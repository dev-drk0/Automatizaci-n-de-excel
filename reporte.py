from __future__ import annotations

import re
from datetime import datetime
from pathlib import Path
from typing import Dict, Any, List
import pandas as pd
from pypdf import PdfReader

# Librerías para ponerle esteroides visuales a Excel
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.chart import BarChart, Reference


def extract_invoice_data(pdf_path: Path) -> Dict[str, Any]:
    """
    Analiza a fondo el PDF de Invoice Generator.
    Extrae de forma robusta: Proveedor, Fecha, Cantidad de Montos Detectados y Total Real.
    """
    reader = PdfReader(pdf_path)
    full_text = ""
    for page in reader.pages:
        text = page.extract_text()
        if text:
            full_text += text + "\n"

    # --- 1. PROVEEDOR ---
    clean_filename = pdf_path.stem.replace("_", " ").replace("-", " ").title()
    lines = [line.strip() for line in full_text.split("\n") if line.strip()]
    vendor = "No detectado"
    if lines:
        for line in lines[:4]:
            if not any(kw in line.lower() for kw in ["invoice", "factura", "fecha", "date", "bill to", "remit to"]):
                vendor = line
                break
        if vendor == "No detectado":
            vendor = clean_filename

    # --- 2. FECHA ---
    date_val = "No detectada"
    numeric_date = re.search(r"(\b\d{1,2}[/-]\d{1,2}[/-]\d{2,4}\b)|(\b\d{4}[/-]\d{1,2}[/-]\d{1,2}\b)", full_text)
    if numeric_date:
        date_val = numeric_date.group(0)
    else:
        text_date = re.search(r"(\b\d{1,2}\s+de\s+[a-zA-Z]+\s+de\s+\d{4}\b)|(\b[a-zA-Z]+\s+\d{1,2},?\s+\d{4}\b)", full_text)
        if text_date:
            date_val = text_date.group(0)

    # --- 3. MONTOS ---
    all_amounts = re.findall(r"[\$-]?\s*([0-9]{1,3}(?:,[0-9]{3})*(?:\.[0-9]{2}))", full_text)
    valid_amounts = []
    for amt in all_amounts:
        try:
            val = float(amt.replace(",", ""))
            if 0.5 <= val < 999999.0:
                valid_amounts.append(val)
        except ValueError:
            continue

    unique_amounts = list(dict.fromkeys(valid_amounts))
    total_amount = 0.0
    items_count = 0

    if unique_amounts:
        total_amount = max(unique_amounts)
        concept_amounts = [a for a in unique_amounts if a < total_amount]
        items_count = len(concept_amounts) if concept_amounts else 1

    category = "Gastos Operativos"
    if "cloud" in full_text.lower() or "aws" in full_text.lower() or "software" in full_text.lower():
        category = "Tecnología / Software"
    elif "marketing" in full_text.lower() or "ads" in full_text.lower():
        category = "Publicidad"

    return {
        "Nombre del Archivo": pdf_path.name,
        "Identificador Visual": clean_filename,
        "Proveedor": vendor,
        "Fecha de Emisión": date_val,
        "Categoría de Gasto": category,
        "Conceptos": items_count,
        "Monto Total": total_amount,
        "Estado": "Revisado / Pendiente"
    }


def process_multiple_files(file_paths: List[str], output_dir: str | Path | None = None) -> Dict[str, Any]:
    base_dir = Path(__file__).resolve().parent
    output_dir = Path(output_dir or "output")
    output_dir.mkdir(parents=True, exist_ok=True)

    records = []
    for p in file_paths:
        path_obj = Path(p)
        if path_obj.suffix.lower() == ".pdf":
            try:
                records.append(extract_invoice_data(path_obj))
            except Exception:
                continue

    if not records:
        raise ValueError("No se pudo extraer información válida de ningún archivo.")

    df_result = pd.DataFrame(records)

    # --- NOMBRE DINÁMICO CON TIMESTAMP ---
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    consolidated_path = output_dir / f"control_facturas_{timestamp}.xlsx"

    # Guardar usando openpyxl como motor para poder editar el diseño
    with pd.ExcelWriter(consolidated_path, engine="openpyxl") as writer:
        df_result.to_excel(writer, index=False, sheet_name="Reporte Consolidado")
        
        # Conseguir acceso a la hoja nativa
        workbook = writer.book
        worksheet = writer.sheets["Reporte Consolidado"]
        
        # --- ESTILOS VISUALES ---
        header_fill = PatternFill(start_color="4C78A8", end_color="4C78A8", fill_type="solid")
        header_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
        data_font = Font(name="Calibri", size=11, color="333333")
        total_fill = PatternFill(start_color="E6EDF5", end_color="E6EDF5", fill_type="solid")
        total_font = Font(name="Calibri", size=11, bold=True, color="000000")
        
        thin_side = Side(border_style="thin", color="CCCCCC")
        data_border = Border(left=thin_side, right=thin_side, top=thin_side, bottom=thin_side)
        
        # Formatear encabezados
        for cell in worksheet[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center")
            
        # Formatear filas de datos
        num_rows = len(df_result)
        for row in worksheet.iter_rows(min_row=2, max_row=num_rows + 1):
            for cell in row:
                cell.font = data_font
                cell.border = data_border
                # Si es la columna de Monto Total (Columna G / 7), darle formato de moneda
                if cell.column == 7:
                    cell.number_format = "$#,##0.00"
                    cell.alignment = Alignment(horizontal="right")

        # --- FILA DE TOTAL GENERAL AL FINAL ---
        total_row_idx = num_rows + 2
        worksheet.cell(row=total_row_idx, column=1, value="TOTAL GENERAL").font = total_font
        worksheet.cell(row=total_row_idx, column=7, value=f"=SUM(G2:G{total_row_idx-1})").font = total_font
        worksheet.cell(row=total_row_idx, column=7).number_format = "$#,##0.00"
        
        # Darle color gris claro a la fila de totales
        for col_idx in range(1, 9):
            cell = worksheet.cell(row=total_row_idx, column=col_idx)
            cell.fill = total_fill
            if col_idx == 7:
                cell.alignment = Alignment(horizontal="right")

        # --- AUTOAJUSTAR COLUMNAS (Evita el texto apachurrado) ---
        for col in worksheet.columns:
            max_len = 0
            col_letter = get_column_letter(col[0].column)
            for cell in col:
                if cell.value:
                    max_len = max(max_len, len(str(cell.value)))
            worksheet.column_dimensions[col_letter].width = max(max_len + 3, 12)

        # --- AÑADIR GRÁFICA DE BARRAS NATIVA ---
        chart = BarChart()
        chart.type = "col"
        chart.style = 10
        chart.title = "Distribución de Gastos por Factura"
        chart.y_axis.title = "Monto ($)"
        chart.x_axis.title = "Proveedor"
        
        # Datos desde la celda G1 hasta la G(última con datos de facturas)
        data_ref = Reference(worksheet, min_col=7, min_row=1, max_row=total_row_idx-1)
        # Categorías (Eje X) tomadas desde la columna C (Proveedor)
        cats_ref = Reference(worksheet, min_col=3, min_row=2, max_row=total_row_idx-1)
        
        chart.add_data(data_ref, titles_from_data=True)
        chart.set_categories(cats_ref)
        chart.legend = None # No necesitamos leyenda al ser una sola serie
        
        # Posicionar la gráfica dejando espacio a la derecha de la tabla
        worksheet.add_chart(chart, "J2")

    return {
        "status": "success",
        "total_files": len(file_paths),
        "rows_processed": len(df_result),
        "output_file": consolidated_path,
        "grand_total": float(df_result["Monto Total"].sum())
    }